import os
import sqlite3
from collections import defaultdict
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils import get_column_letter

import pytz

from team_manager import get_inactive_ids

# ─── КОНСТАНТЫ НОРМ ───────────────────────────────────────────────────────────
MIN_DOCTORS_DAY = 8
MAX_DOCTORS_DAY = 12
MIN_PHARM_DAY   = 6
MAX_PHARM_DAY   = 10

TZ = pytz.timezone("Asia/Tashkent")
DAYS_RU = {0: "Пн", 1: "Вт", 2: "Ср", 3: "Чт", 4: "Пт", 5: "Сб", 6: "Вс"}

# Категории потенциала
def _category(visits_count: int) -> str:
    if visits_count >= 5:   return "A"
    elif visits_count >= 3: return "B"
    elif visits_count >= 1: return "C"
    return "—"

def _category_fill(cat: str):
    if cat == "A": return _fill("C6EFCE")
    if cat == "B": return _fill("FFEB9C")
    if cat == "C": return _fill("FCE4D6")
    return _fill("FFFFFF")

# ─── СТИЛИ ────────────────────────────────────────────────────────────────────
def _hdr_font():   return Font(name="Arial", bold=True, color="FFFFFF", size=11)
def _body_font():  return Font(name="Arial", size=10)
def _title_font(): return Font(name="Arial", bold=True, size=13, color="1F4E79")
def _fill(hex_color): return PatternFill("solid", start_color=hex_color)

FILL_HEADER  = _fill("1F4E79")
FILL_SUBHEAD = _fill("2E75B6")
FILL_GREEN   = _fill("E2EFDA")
FILL_YELLOW  = _fill("FFEB9C")
FILL_ORANGE  = _fill("FCE4D6")
FILL_RED     = _fill("FF7676")
FILL_BLUE    = _fill("DEEAF1")
FILL_PURPLE  = _fill("E2D0F0")
FILL_ALT     = _fill("F5F5F5")
FILL_TOTAL   = _fill("D6E4F0")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def _border():
    t = Side(style="thin", color="BBBBBB")
    return Border(left=t, right=t, top=t, bottom=t)

def _apply_header(cell, fill=None):
    cell.font = _hdr_font(); cell.fill = fill or FILL_HEADER
    cell.alignment = CENTER; cell.border = _border()

def _apply_body(cell, fill=None, align=CENTER):
    cell.font = _body_font(); cell.fill = fill or _fill("FFFFFF")
    cell.alignment = align; cell.border = _border()

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ─── ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ──────────────────────────────────────────────────

def _strip_tz(dt):
    if dt and hasattr(dt, 'tzinfo') and dt.tzinfo:
        return dt.replace(tzinfo=None)
    return dt

def _parse_dt(s):
    try:
        return datetime.strptime(s, "%d.%m.%Y %H:%M")
    except Exception:
        return None

def _working_days_completed(start_dt, end_dt):
    """Считает только завершённые рабочие дни (сегодня не считается).
    Нормализуем даты до уровня дня, чтобы 02.04 в 00:00:00 считался как завершённый.
    """
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday = today - timedelta(days=1)
    start_day = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    end_day   = min(end_dt, yesterday).replace(hour=0, minute=0, second=0, microsecond=0)
    if end_day < start_day:
        return 0
    return sum(
        1 for d in range((end_day - start_day).days + 1)
        if (start_day + timedelta(days=d)).weekday() < 5
    )

def _norms(working_days):
    return {
        "min_doc":   MIN_DOCTORS_DAY * working_days,
        "max_doc":   MAX_DOCTORS_DAY * working_days,
        "min_ph":    MIN_PHARM_DAY   * working_days,
        "max_ph":    MAX_PHARM_DAY   * working_days,
        "min_total": (MIN_DOCTORS_DAY + MIN_PHARM_DAY) * working_days,
        "max_total": (MAX_DOCTORS_DAY + MAX_PHARM_DAY) * working_days,
    }

def _verdict(total, min_total, max_total):
    if max_total == 0:              return "— нет данных"
    if total < min_total:           return "❌ Норма не выполнена"
    elif total == min_total:        return "✅ Выполнен минимум"
    elif total < max_total:         return "👍 Выполнено средне"
    elif total == max_total:        return "🏆 Выполнен максимум"
    else:                           return "🚀 Выше максимума"

def _verdict_fill(v):
    if "не выполнена" in v: return FILL_RED
    if "минимум"      in v: return FILL_YELLOW
    if "средне"       in v: return FILL_ORANGE
    if "максимум"     in v: return FILL_GREEN
    if "Выше"         in v: return FILL_PURPLE
    return FILL_ALT

def _pct(val, norm):
    return round(val / norm * 100, 1) if norm else 0

def _is_suspicious(visit) -> bool:
    """Визит подозрительный: ФИО — цифра, одиночный символ или пустое."""
    name = (visit[5] or "").strip()
    if not name: return True
    if name.isdigit(): return True
    if len(name) <= 2: return True
    return False

# ─── ЧТЕНИЕ ДАННЫХ ────────────────────────────────────────────────────────────

def fetch_visits_range(visits_db, start_dt, end_dt):
    """Фильтрация в Python — дата в базе хранится как ДД.ММ.ГГГГ ЧЧ:ММ."""
    start_dt = _strip_tz(start_dt)
    end_dt   = _strip_tz(end_dt)
    # Исключаем текущий день
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    if end_dt >= today:
        end_dt = today - timedelta(seconds=1)

    with sqlite3.connect(visits_db) as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM visits ORDER BY telegram_id, date_time")
        rows = cur.fetchall()

    result = []
    for row in rows:
        dt = _parse_dt(row[4])
        if dt and start_dt <= dt <= end_dt:
            result.append(row)
    return result

def fetch_all_visits_history(visits_db):
    """Все визиты из базы без фильтра по дате (для базы врачей/аптек)."""
    with sqlite3.connect(visits_db) as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM visits ORDER BY telegram_id, date_time")
        return cur.fetchall()

def fetch_all_users(users_db, active_only=True) -> list:
    inactive = get_inactive_ids() if active_only else []
    with sqlite3.connect(users_db) as conn:
        cur = conn.cursor()
        cur.execute("SELECT telegram_id, full_name, region, birth_date, phone_number FROM users ORDER BY full_name")
        rows = cur.fetchall()
    if active_only:
        return [r for r in rows if r[0] not in inactive]
    return rows

# ─── ЛИСТ 1: ОСНОВНАЯ ТАБЛИЦА ─────────────────────────────────────────────────

def _sheet_main(wb, users, visits_by_user, start_dt, end_dt, period_label):
    ws = wb.active
    ws.title = "📊 Аналитика"

    ws.merge_cells("A1:O1")
    ws["A1"] = f"Аналитика команды | {period_label} | {start_dt.strftime('%d.%m.%Y')} — {end_dt.strftime('%d.%m.%Y')} (текущий день не учитывается)"
    ws["A1"].font = _title_font(); ws["A1"].alignment = CENTER; ws["A1"].fill = FILL_BLUE
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:O2")
    ws["A2"] = f"Нормы: Врачи {MIN_DOCTORS_DAY}–{MAX_DOCTORS_DAY}/день  |  Аптеки/Оптом {MIN_PHARM_DAY}–{MAX_PHARM_DAY}/день  |  Учитываются только завершённые рабочие дни"
    ws["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")
    ws["A2"].alignment = LEFT; ws.row_dimensions[2].height = 18

    headers = [
        "№", "Сотрудник", "Регион",
        "Врачи", "Норма врачи (мин)", "% врачи",
        "Аптеки", "Норма аптеки (мин)", "% аптеки",
        "Оптом", "Итого визитов", "Норма итого (мин)",
        "% итого", "Вердикт", "Раб. дней"
    ]
    ws.append(headers)
    for cell in ws[3]: _apply_header(cell)
    ws.row_dimensions[3].height = 40

    for idx, user in enumerate(users, 1):
        uid    = user[0]
        visits = [v for v in visits_by_user.get(uid, []) if not _is_suspicious(v)]
        absent = len([v for v in visits if v[3] == "Не вышел"])

        # Рабочие дни только завершённые
        wd = _working_days_completed(start_dt, end_dt)
        wd = max(wd - absent, 1) if wd > 0 else 1
        n  = _norms(wd)

        doctors = len([v for v in visits if v[3] == "🩺 Врач"])
        pharm   = len([v for v in visits if v[3] == "💊 Аптека"])
        dist    = len([v for v in visits if v[3] == "🚚 Дистрибьютор"])
        total   = doctors + pharm + dist
        verd    = _verdict(total, n["min_total"], n["max_total"])

        ws.append([
            idx, user[1], user[2],
            doctors, n["min_doc"], f"{_pct(doctors, n['min_doc'])}%",
            pharm,   n["min_ph"],  f"{_pct(pharm, n['min_ph'])}%",
            dist, total, n["min_total"], f"{_pct(total, n['min_total'])}%",
            verd, wd
        ])

        vf   = _verdict_fill(verd)
        alt  = _fill("F5F5F5") if idx % 2 == 0 else _fill("FFFFFF")
        for ci, cell in enumerate(ws[idx + 3], 1):
            if ci == 14:   _apply_body(cell, vf, LEFT)
            elif ci in (6, 9, 13):
                try:
                    pv = float(str(cell.value).replace("%", ""))
                    pf = FILL_GREEN if pv >= 100 else FILL_YELLOW if pv >= 60 else FILL_RED
                except Exception: pf = alt
                _apply_body(cell, pf)
            else: _apply_body(cell, alt)

    ws.freeze_panes = "A4"
    _set_col_widths(ws, [4, 22, 14, 8, 18, 10, 8, 18, 10, 8, 12, 18, 10, 26, 10])

# ─── ЛИСТ 2: АКТИВНОСТЬ ПО ДНЯМ НЕДЕЛИ ───────────────────────────────────────

def _sheet_weekdays(wb, visits_by_user, users, start_dt, end_dt):
    ws = wb.create_sheet("📅 По дням недели")
    user_map = {u[0]: u[1] for u in users}

    ws.merge_cells("A1:K1")
    ws["A1"] = "Активность по дням недели (только завершённые дни, без подозрительных визитов)"
    ws["A1"].font = _title_font(); ws["A1"].alignment = CENTER; ws["A1"].fill = FILL_BLUE

    headers = ["Сотрудник", "Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Итого", "Норма (мин)", "% выполн.", "Вердикт"]
    ws.append(headers)
    for cell in ws[2]: _apply_header(cell)

    wd_total = _working_days_completed(start_dt, end_dt)

    for idx, (uid, visits) in enumerate(visits_by_user.items(), 3):
        clean   = [v for v in visits if not _is_suspicious(v) and v[3] != "Не вышел"]
        absent  = len([v for v in visits if v[3] == "Не вышел"])
        wd      = max(wd_total - absent, 1) if wd_total > 0 else 1
        n       = _norms(wd)
        norm_min = n["min_total"]

        day_counts = defaultdict(int)
        for v in clean:
            dt = _parse_dt(v[4])
            if dt: day_counts[dt.weekday()] += 1

        total = sum(day_counts.values())
        pct   = _pct(total, norm_min)
        verd  = _verdict(total, norm_min, n["max_total"])

        row = [user_map.get(uid, str(uid))]
        for d in range(6):
            row.append(day_counts.get(d, 0))
        row += [total, norm_min, f"{pct}%", verd]
        ws.append(row)

        vf  = _verdict_fill(verd)
        alt = _fill("F5F5F5") if idx % 2 == 0 else _fill("FFFFFF")
        for ci, cell in enumerate(ws[idx], 1):
            if ci == 11:   _apply_body(cell, vf, LEFT)
            elif ci == 10:
                try:
                    pv = float(str(cell.value).replace("%", ""))
                    pf = FILL_GREEN if pv >= 100 else FILL_YELLOW if pv >= 60 else FILL_RED
                except Exception: pf = alt
                _apply_body(cell, pf)
            else: _apply_body(cell, alt)

    _set_col_widths(ws, [22, 6, 6, 6, 6, 6, 6, 8, 12, 12, 26])
    ws.freeze_panes = "A3"

    # График
    if len(visits_by_user) > 0:
        chart = BarChart()
        chart.type = "col"; chart.title = "Визиты по дням недели"
        chart.y_axis.title = "Кол-во визитов"; chart.style = 10
        chart.width = 20; chart.height = 14
        data = Reference(ws, min_col=2, max_col=7, min_row=2, max_row=2 + len(visits_by_user))
        cats = Reference(ws, min_col=1, min_row=3, max_row=2 + len(visits_by_user))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "M2")

# ─── ЛИСТ 3: ТОП ОРГАНИЗАЦИЙ ПО РЕГИОНАМ ─────────────────────────────────────

def _sheet_orgs(wb, all_visits, users):
    ws = wb.create_sheet("🏆 Орг-ции по регионам")
    user_map = {u[0]: u[1] for u in users}

    # Собираем данные: регион → организация → {визитов, сотрудники}
    region_org = defaultdict(lambda: defaultdict(lambda: {"count": 0, "employees": set(), "cats": set()}))
    for v in all_visits:
        if v[3] == "Не вышел" or not v[6] or _is_suspicious(v): continue
        region  = user_map.get(v[1], v[2] or "—")
        # Регион берём из профиля пользователя
        for u in users:
            if u[0] == v[1]:
                region = u[2]
                break
        org = v[6]
        region_org[region][org]["count"]     += 1
        region_org[region][org]["employees"].add(user_map.get(v[1], "—"))
        region_org[region][org]["cats"].add(v[3])

    row_idx = 1
    for region, orgs in sorted(region_org.items()):
        sorted_orgs = sorted(orgs.items(), key=lambda x: x[1]["count"], reverse=True)
        top     = sorted_orgs[:10]
        bottom  = sorted_orgs[-5:] if len(sorted_orgs) > 10 else []

        # Заголовок региона
        ws.merge_cells(f"A{row_idx}:F{row_idx}")
        ws[f"A{row_idx}"] = f"📍 Регион: {region}"
        ws[f"A{row_idx}"].font = _title_font()
        ws[f"A{row_idx}"].fill = FILL_BLUE
        ws[f"A{row_idx}"].alignment = LEFT
        row_idx += 1

        # Топ
        ws.merge_cells(f"A{row_idx}:F{row_idx}")
        ws[f"A{row_idx}"] = "🏆 Топ-10 организаций"
        ws[f"A{row_idx}"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        ws[f"A{row_idx}"].fill = FILL_SUBHEAD
        ws[f"A{row_idx}"].alignment = CENTER
        row_idx += 1

        hdr_row = row_idx
        for col, h in enumerate(["№", "Организация", "Визитов", "Категории", "Сотрудники", "Потенциал"], 1):
            cell = ws.cell(row=hdr_row, column=col, value=h)
            _apply_header(cell, FILL_SUBHEAD)
        row_idx += 1

        for i, (org, data) in enumerate(top, 1):
            cat  = _category(data["count"])
            emps = ", ".join(sorted(data["employees"]))
            cats = ", ".join(data["cats"])
            row_fill = FILL_GREEN if i <= 3 else (_fill("FFFFFF") if i % 2 else FILL_ALT)
            vals = [i, org, data["count"], cats, emps, cat]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                if col == 6: _apply_body(cell, _category_fill(cat))
                elif col == 2: _apply_body(cell, row_fill, LEFT)
                else: _apply_body(cell, row_fill)
            row_idx += 1

        if bottom:
            # Аутсайдеры
            ws.merge_cells(f"A{row_idx}:F{row_idx}")
            ws[f"A{row_idx}"] = "⚠️ Аутсайдеры"
            ws[f"A{row_idx}"].font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
            ws[f"A{row_idx}"].fill = _fill("C55A11")
            ws[f"A{row_idx}"].alignment = CENTER
            row_idx += 1

            for col, h in enumerate(["№", "Организация", "Визитов", "Категории", "Сотрудники", "Потенциал"], 1):
                cell = ws.cell(row=row_idx, column=col, value=h)
                _apply_header(cell, FILL_SUBHEAD)
            row_idx += 1

            for i, (org, data) in enumerate(reversed(bottom), 1):
                cat  = _category(data["count"])
                emps = ", ".join(sorted(data["employees"]))
                cats = ", ".join(data["cats"])
                vals = [i, org, data["count"], cats, emps, cat]
                for col, val in enumerate(vals, 1):
                    cell = ws.cell(row=row_idx, column=col, value=val)
                    rf = FILL_RED if data["count"] == 1 else FILL_ORANGE
                    if col == 6: _apply_body(cell, _category_fill(cat))
                    elif col == 2: _apply_body(cell, rf, LEFT)
                    else: _apply_body(cell, rf)
                row_idx += 1

        row_idx += 1  # Пустая строка между регионами

    _set_col_widths(ws, [4, 35, 10, 20, 30, 10])

# ─── ЛИСТ 4: ПО РЕГИОНАМ ──────────────────────────────────────────────────────

def _sheet_regions(wb, all_visits, users):
    ws = wb.create_sheet("📍 Регионы")
    user_map = {u[0]: u[2] for u in users}

    region_data = defaultdict(lambda: defaultdict(int))
    for v in all_visits:
        if v[3] == "Не вышел" or _is_suspicious(v): continue
        region = user_map.get(v[1], v[2] or "—")
        region_data[region][v[3]] += 1

    ws.merge_cells("A1:F1")
    ws["A1"] = "Распределение визитов по регионам"
    ws["A1"].font = _title_font(); ws["A1"].alignment = CENTER; ws["A1"].fill = FILL_BLUE

    ws.append(["Регион", "🩺 Врачи", "💊 Аптеки", "🚚 Оптом", "Итого", "% от всех"])
    for cell in ws[2]: _apply_header(cell)

    grand_total = sum(sum(c.values()) for c in region_data.values())
    totals = {"doc": 0, "ph": 0, "di": 0, "tot": 0}

    for idx, (region, cats) in enumerate(sorted(region_data.items()), 3):
        doc = cats.get("🩺 Врач", 0)
        ph  = cats.get("💊 Аптека", 0)
        di  = cats.get("🚚 Дистрибьютор", 0)
        tot = doc + ph + di
        totals["doc"] += doc; totals["ph"] += ph
        totals["di"]  += di;  totals["tot"] += tot

        ws.append([region, doc, ph, di, tot, f"{_pct(tot, grand_total)}%"])
        fill = FILL_ALT if idx % 2 == 0 else _fill("FFFFFF")
        for cell in ws[idx]: _apply_body(cell, fill)

    # Итоговая строка
    last = ws.max_row + 1
    ws.append(["ОБЩЕЕ:", totals["doc"], totals["ph"], totals["di"], totals["tot"], "100%"])
    for cell in ws[last]:
        cell.font = Font(name="Arial", bold=True, size=11)
        cell.fill = FILL_TOTAL; cell.alignment = CENTER; cell.border = _border()

    _set_col_widths(ws, [20, 10, 10, 10, 10, 10])
    ws.freeze_panes = "A3"

    # Круговая диаграмма
    cat_totals = defaultdict(int)
    for v in all_visits:
        if v[3] != "Не вышел" and not _is_suspicious(v):
            cat_totals[v[3]] += 1

    pie_start = ws.max_row + 2
    ws.cell(row=pie_start, column=1, value="Категория").font = Font(name="Arial", bold=True)
    ws.cell(row=pie_start, column=2, value="Кол-во").font   = Font(name="Arial", bold=True)
    for i, (cat, cnt) in enumerate(cat_totals.items(), pie_start + 1):
        ws.cell(row=i, column=1, value=cat)
        ws.cell(row=i, column=2, value=cnt)

    pie = PieChart()
    pie.title = "По категориям"; pie.style = 10; pie.width = 16; pie.height = 14
    data = Reference(ws, min_col=2, min_row=pie_start, max_row=pie_start + len(cat_totals))
    cats_ref = Reference(ws, min_col=1, min_row=pie_start + 1, max_row=pie_start + len(cat_totals))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats_ref)
    ws.add_chart(pie, "H2")

# ─── ЛИСТ 5: ДИНАМИКА ─────────────────────────────────────────────────────────

def _sheet_trend(wb, all_visits, start_dt, end_dt):
    ws = wb.create_sheet("📈 Динамика")

    ws.merge_cells("A1:C1")
    ws["A1"] = "Динамика визитов по дням (текущий день не учитывается)"
    ws["A1"].font = _title_font(); ws["A1"].alignment = CENTER; ws["A1"].fill = FILL_BLUE

    ws.append(["Дата", "День недели", "Визитов"])
    for cell in ws[2]: _apply_header(cell)

    day_counts = defaultdict(int)
    for v in all_visits:
        if v[3] == "Не вышел" or _is_suspicious(v): continue
        dt = _parse_dt(v[4])
        if dt: day_counts[dt.strftime("%d.%m.%Y")] += 1

    today   = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    current = start_dt
    row_idx = 3
    while current < today and current <= end_dt:
        date_str = current.strftime("%d.%m.%Y")
        ws.append([date_str, DAYS_RU[current.weekday()], day_counts.get(date_str, 0)])
        fill = _fill("F0F0F0") if current.weekday() >= 5 else (FILL_ALT if row_idx % 2 == 0 else _fill("FFFFFF"))
        for cell in ws[row_idx]: _apply_body(cell, fill)
        current  += timedelta(days=1)
        row_idx  += 1

    _set_col_widths(ws, [14, 12, 12])
    ws.freeze_panes = "A3"

    line = LineChart()
    line.title = "Динамика визитов"; line.style = 10
    line.y_axis.title = "Кол-во"; line.x_axis.title = "Дата"
    line.width = 24; line.height = 14
    data = Reference(ws, min_col=3, min_row=2, max_row=row_idx - 1)
    cats = Reference(ws, min_col=1, min_row=3, max_row=row_idx - 1)
    line.add_data(data, titles_from_data=True)
    line.set_categories(cats)
    ws.add_chart(line, "E2")

# ─── ЛИСТ 6: ВРАЧИ ────────────────────────────────────────────────────────────

def _sheet_doctors(wb, all_visits, users):
    ws = wb.create_sheet("👨‍⚕️ Врачи")
    user_map = {u[0]: (u[1], u[2]) for u in users}

    doctor_data = defaultdict(lambda: {"visits": 0, "orgs": set(), "specs": set(), "regions": set(), "employees": set()})
    for v in all_visits:
        if v[3] != "🩺 Врач" or not v[5] or _is_suspicious(v): continue
        name = v[5]
        doctor_data[name]["visits"]    += 1
        if v[6]: doctor_data[name]["orgs"].add(v[6])
        if v[7]: doctor_data[name]["specs"].add(v[7])
        emp = user_map.get(v[1])
        if emp:
            doctor_data[name]["regions"].add(emp[1])
            doctor_data[name]["employees"].add(emp[0])

    sorted_docs = sorted(doctor_data.items(), key=lambda x: x[1]["visits"], reverse=True)
    top_docs    = sorted_docs[:15]
    bottom_docs = sorted_docs[-10:] if len(sorted_docs) > 15 else []

    headers = ["№", "ФИО врача", "Специализация", "Организация", "Регион", "Сотрудник", "Визитов", "Потенциал"]

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"] = "Топ-15 врачей"
    ws["A1"].font = _title_font(); ws["A1"].alignment = CENTER; ws["A1"].fill = FILL_GREEN

    ws.append(headers)
    for cell in ws[2]: _apply_header(cell, FILL_SUBHEAD)

    for i, (name, data) in enumerate(top_docs, 1):
        cat  = _category(data["visits"])
        fill = FILL_GREEN if i <= 3 else (_fill("FFFFFF") if i % 2 else FILL_ALT)
        ws.append([
            i, name,
            ", ".join(data["specs"]),
            ", ".join(data["orgs"]),
            ", ".join(data["regions"]),
            ", ".join(data["employees"]),
            data["visits"], cat
        ])
        for ci, cell in enumerate(ws[i + 2], 1):
            if ci == 8: _apply_body(cell, _category_fill(cat))
            elif ci in (2, 3, 4, 5, 6): _apply_body(cell, fill, LEFT)
            else: _apply_body(cell, fill)

    gap = ws.max_row + 2
    ws.merge_cells(f"A{gap}:{get_column_letter(len(headers))}{gap}")
    ws[f"A{gap}"] = "Аутсайдеры — врачи с минимальным числом визитов"
    ws[f"A{gap}"].font = _title_font(); ws[f"A{gap}"].fill = FILL_ORANGE; ws[f"A{gap}"].alignment = CENTER

    hdr = gap + 1
    ws.append(headers)
    for cell in ws[hdr]: _apply_header(cell, FILL_SUBHEAD)

    for i, (name, data) in enumerate(reversed(bottom_docs), 1):
        cat  = _category(data["visits"])
        fill = FILL_RED if data["visits"] == 1 else FILL_ORANGE
        ws.append([
            i, name,
            ", ".join(data["specs"]),
            ", ".join(data["orgs"]),
            ", ".join(data["regions"]),
            ", ".join(data["employees"]),
            data["visits"], cat
        ])
        for ci, cell in enumerate(ws[ws.max_row], 1):
            if ci == 8: _apply_body(cell, _category_fill(cat))
            elif ci in (2, 3, 4, 5, 6): _apply_body(cell, fill, LEFT)
            else: _apply_body(cell, fill)

    _set_col_widths(ws, [4, 28, 22, 28, 16, 22, 10, 12])

# ─── ЛИСТ 7: БАЗА КЛИЕНТОВ ────────────────────────────────────────────────────

def _sheet_client_base(wb, all_visits_history, users):
    user_map = {u[0]: (u[1], u[2]) for u in users}

    # ── База врачей ──────────────────────────────────────────────────────────
    ws_doc = wb.create_sheet("🗂 База врачей")
    doc_data = defaultdict(lambda: {"visits": 0, "orgs": set(), "specs": set(), "regions": set(), "employees": set()})
    for v in all_visits_history:
        if v[3] != "🩺 Врач" or not v[5] or _is_suspicious(v): continue
        name = v[5]
        doc_data[name]["visits"] += 1
        if v[6]: doc_data[name]["orgs"].add(v[6])
        if v[7]: doc_data[name]["specs"].add(v[7])
        emp = user_map.get(v[1])
        if emp:
            doc_data[name]["regions"].add(emp[1])
            doc_data[name]["employees"].add(emp[0])

    ws_doc.merge_cells("A1:H1")
    ws_doc["A1"] = "База врачей (вся история) | A = 5+ визитов, B = 3-4, C = 1-2"
    ws_doc["A1"].font = _title_font(); ws_doc["A1"].alignment = CENTER; ws_doc["A1"].fill = FILL_BLUE

    doc_headers = ["№", "ФИО врача", "Специализация", "Организация", "Регион", "Сотрудник", "Всего визитов", "Категория"]
    ws_doc.append(doc_headers)
    for cell in ws_doc[2]: _apply_header(cell)

    for i, (name, data) in enumerate(sorted(doc_data.items(), key=lambda x: x[1]["visits"], reverse=True), 1):
        cat  = _category(data["visits"])
        fill = _fill("FFFFFF") if i % 2 else FILL_ALT
        ws_doc.append([
            i, name,
            ", ".join(data["specs"]),
            ", ".join(data["orgs"]),
            ", ".join(data["regions"]),
            ", ".join(data["employees"]),
            data["visits"], cat
        ])
        for ci, cell in enumerate(ws_doc[i + 2], 1):
            if ci == 8: _apply_body(cell, _category_fill(cat))
            elif ci in (2, 3, 4, 5, 6): _apply_body(cell, fill, LEFT)
            else: _apply_body(cell, fill)

    _set_col_widths(ws_doc, [4, 28, 22, 28, 16, 22, 14, 12])

    # ── База аптек ───────────────────────────────────────────────────────────
    ws_ph = wb.create_sheet("🗂 База аптек")
    ph_data = defaultdict(lambda: {"visits": 0, "regions": set(), "employees": set()})
    for v in all_visits_history:
        if v[3] != "💊 Аптека" or not v[6] or _is_suspicious(v): continue
        org = v[6]
        ph_data[org]["visits"] += 1
        emp = user_map.get(v[1])
        if emp:
            ph_data[org]["regions"].add(emp[1])
            ph_data[org]["employees"].add(emp[0])

    ws_ph.merge_cells("A1:G1")
    ws_ph["A1"] = "База аптек (вся история) | A = 5+ визитов, B = 3-4, C = 1-2"
    ws_ph["A1"].font = _title_font(); ws_ph["A1"].alignment = CENTER; ws_ph["A1"].fill = FILL_BLUE

    ph_headers = ["№", "Название аптеки", "Регион", "Сотрудник", "Всего визитов", "Категория", "Специализация/Должность"]
    ws_ph.append(ph_headers)
    for cell in ws_ph[2]: _apply_header(cell)

    for i, (org, data) in enumerate(sorted(ph_data.items(), key=lambda x: x[1]["visits"], reverse=True), 1):
        cat  = _category(data["visits"])
        fill = _fill("FFFFFF") if i % 2 else FILL_ALT
        # Ищем специализацию из визитов
        specs = set()
        for v in all_visits_history:
            if v[3] == "💊 Аптека" and v[6] == org and v[7]: specs.add(v[7])
        ws_ph.append([
            i, org,
            ", ".join(data["regions"]),
            ", ".join(data["employees"]),
            data["visits"], cat,
            ", ".join(specs)
        ])
        for ci, cell in enumerate(ws_ph[i + 2], 1):
            if ci == 6: _apply_body(cell, _category_fill(cat))
            elif ci in (2, 3, 4, 7): _apply_body(cell, fill, LEFT)
            else: _apply_body(cell, fill)

    _set_col_widths(ws_ph, [4, 30, 16, 22, 14, 12, 22])

    # ── База оптом ───────────────────────────────────────────────────────────
    ws_di = wb.create_sheet("🗂 База оптом")
    di_data = defaultdict(lambda: {"visits": 0, "regions": set(), "employees": set()})
    for v in all_visits_history:
        if v[3] != "🚚 Дистрибьютор" or not v[6] or _is_suspicious(v): continue
        org = v[6]
        di_data[org]["visits"] += 1
        emp = user_map.get(v[1])
        if emp:
            di_data[org]["regions"].add(emp[1])
            di_data[org]["employees"].add(emp[0])

    ws_di.merge_cells("A1:G1")
    ws_di["A1"] = "База оптом/дистрибьюторов (вся история) | A = 5+ визитов, B = 3-4, C = 1-2"
    ws_di["A1"].font = _title_font(); ws_di["A1"].alignment = CENTER; ws_di["A1"].fill = FILL_BLUE

    di_headers = ["№", "Название организации", "Регион", "Сотрудник", "Всего визитов", "Категория", "Должность"]
    ws_di.append(di_headers)
    for cell in ws_di[2]: _apply_header(cell)

    for i, (org, data) in enumerate(sorted(di_data.items(), key=lambda x: x[1]["visits"], reverse=True), 1):
        cat  = _category(data["visits"])
        fill = _fill("FFFFFF") if i % 2 else FILL_ALT
        specs = set()
        for v in all_visits_history:
            if v[3] == "🚚 Дистрибьютор" and v[6] == org and v[7]: specs.add(v[7])
        ws_di.append([
            i, org,
            ", ".join(data["regions"]),
            ", ".join(data["employees"]),
            data["visits"], cat,
            ", ".join(specs)
        ])
        for ci, cell in enumerate(ws_di[i + 2], 1):
            if ci == 6: _apply_body(cell, _category_fill(cat))
            elif ci in (2, 3, 4, 7): _apply_body(cell, fill, LEFT)
            else: _apply_body(cell, fill)

    _set_col_widths(ws_di, [4, 30, 16, 22, 14, 12, 22])

# ─── ГЛАВНАЯ ФУНКЦИЯ ──────────────────────────────────────────────────────────

def generate_analytics_excel(visits_db, users_db, start_dt, end_dt, period_name, filename):
    start_dt = _strip_tz(start_dt)
    end_dt   = _strip_tz(end_dt)

    # Текущий день не учитывается
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    if end_dt >= today:
        end_dt = today - timedelta(seconds=1)

    users      = fetch_all_users(users_db, active_only=True)
    all_visits = fetch_visits_range(visits_db, start_dt, end_dt)
    all_history= fetch_all_visits_history(visits_db)

    visits_by_user = defaultdict(list)
    for v in all_visits:
        visits_by_user[v[1]].append(v)

    wb = openpyxl.Workbook()

    _sheet_main(wb, users, visits_by_user, start_dt, end_dt, period_name)
    _sheet_weekdays(wb, visits_by_user, users, start_dt, end_dt)
    _sheet_orgs(wb, all_visits, users)
    _sheet_regions(wb, all_visits, users)
    _sheet_trend(wb, all_visits, start_dt, end_dt)
    _sheet_doctors(wb, all_visits, users)
    _sheet_client_base(wb, all_history, users)

    wb.save(filename)
    return filename
