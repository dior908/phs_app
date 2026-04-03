import os
import sqlite3
from collections import defaultdict
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pytz
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import threading

from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler, filters,
    ConversationHandler, ContextTypes
)

from analytics import generate_analytics_excel
from team_manager import (
    get_all_users_with_status, get_active_users,
    deactivate_employee, activate_employee,
    get_first_visit_date, get_days_worked
)

# ─── КОНФИГ ───────────────────────────────────────────────────────────────────
ADMIN_BOT_TOKEN = "8341249968:AAG55xcFQdl-54d30fvE9OLVX-bhI8FiUsU"  # ← замени токен
ALLOWED_IDS     = [6669377232]

MAIN_BOT_DIR = os.path.dirname(os.path.abspath(__file__))
USERS_DB     = os.path.join(MAIN_BOT_DIR, "databases", "users.db")
VISITS_DB    = os.path.join(MAIN_BOT_DIR, "databases", "visits.db")
TOURPLAN_DB  = os.path.join(MAIN_BOT_DIR, "databases", "tourplan.db")
REPORTS_DIR  = os.path.join(MAIN_BOT_DIR, "admin_reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

FONTS_DIR  = os.path.join(MAIN_BOT_DIR, "fonts")
NOTO_SANS  = os.path.join(FONTS_DIR, "NotoSans-Regular.ttf")
NOTO_EMOJI = os.path.join(FONTS_DIR, "NotoEmoji-Regular.ttf")
if os.path.exists(NOTO_SANS) and os.path.exists(NOTO_EMOJI):
    pdfmetrics.registerFont(TTFont("NotoSans",  NOTO_SANS))
    pdfmetrics.registerFont(TTFont("NotoEmoji", NOTO_EMOJI))

TZ = pytz.timezone("Asia/Tashkent")

# ─── СОСТОЯНИЯ ────────────────────────────────────────────────────────────────
(
    MAIN_MENU,
    CHOOSE_EMPLOYEE, CHOOSE_PERIOD, CHOOSE_START, CHOOSE_END,
    SUMMARY_PERIOD, SUMMARY_START, SUMMARY_END,
    EXCEL_PERIOD, EXCEL_START, EXCEL_END,
    TOURPLAN_EMPLOYEE,
    ANALYTICS_PERIOD, ANALYTICS_START, ANALYTICS_END,
    MGMT_MENU,
    TEAM_CHOOSE_EMPLOYEE, TEAM_ACTION,
    TEAM_REPORT_PERIOD, TEAM_REPORT_START, TEAM_REPORT_END,
) = range(21)

# ─── ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ──────────────────────────────────────────────────

def is_allowed(uid): return uid in ALLOWED_IDS

def _get_all_users_db():
    with sqlite3.connect(USERS_DB) as conn:
        cur = conn.cursor()
        cur.execute("SELECT telegram_id, full_name, region, phone_number FROM users ORDER BY full_name")
        return cur.fetchall()

def get_visits_for_user(telegram_id, start_dt, end_dt):
    if hasattr(start_dt, 'tzinfo') and start_dt.tzinfo: start_dt = start_dt.replace(tzinfo=None)
    if hasattr(end_dt,   'tzinfo') and end_dt.tzinfo:   end_dt   = end_dt.replace(tzinfo=None)
    # Исключаем текущий день
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    if end_dt >= today: end_dt = today - timedelta(seconds=1)
    with sqlite3.connect(VISITS_DB) as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM visits WHERE telegram_id = ?", (telegram_id,))
        rows = cur.fetchall()
    result = []
    for row in rows:
        try:
            dt = datetime.strptime(row[4], "%d.%m.%Y %H:%M")
            if start_dt <= dt <= end_dt: result.append(row)
        except Exception: continue
    return result

def get_all_visits_filtered(start_dt, end_dt):
    if hasattr(start_dt, 'tzinfo') and start_dt.tzinfo: start_dt = start_dt.replace(tzinfo=None)
    if hasattr(end_dt,   'tzinfo') and end_dt.tzinfo:   end_dt   = end_dt.replace(tzinfo=None)
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    if end_dt >= today: end_dt = today - timedelta(seconds=1)
    with sqlite3.connect(VISITS_DB) as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM visits ORDER BY telegram_id, date_time")
        rows = cur.fetchall()
    result = []
    for row in rows:
        try:
            dt = datetime.strptime(row[4], "%d.%m.%Y %H:%M")
            if start_dt <= dt <= end_dt: result.append(row)
        except Exception: continue
    return result

def get_tourplan(telegram_id):
    with sqlite3.connect(TOURPLAN_DB) as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM tourplan WHERE telegram_id = ? ORDER BY date", (telegram_id,))
        return cur.fetchall()

def period_bounds(period):
    today = datetime.now(TZ)
    if period == "🗓 За текущий день":
        s = today.replace(hour=0, minute=1, second=0, microsecond=0)
        return s, today, "день"
    elif period == "📆 За текущую неделю":
        s = (today - timedelta(days=today.weekday())).replace(hour=0, minute=1, second=0, microsecond=0)
        return s, today, "неделю"
    elif period == "📊 За текущий месяц":
        s = today.replace(day=1, hour=0, minute=1, second=0, microsecond=0)
        return s, today, "месяц"
    return None, None, None

def auto_delete(path, seconds=3600):
    threading.Timer(seconds, lambda: os.path.exists(path) and os.remove(path)).start()

def _working_days_completed(start_dt, end_dt):
    today     = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    yesterday = today - timedelta(days=1)
    start_day = start_dt.replace(hour=0, minute=0, second=0, microsecond=0)
    end_day   = min(end_dt, yesterday).replace(hour=0, minute=0, second=0, microsecond=0)
    if end_day < start_day: return 0
    return sum(1 for d in range((end_day - start_day).days + 1) if (start_day + timedelta(days=d)).weekday() < 5)

def _pct(val, norm): return round(val / norm * 100, 1) if norm else 0

def _verdict_short(total, min_t, max_t):
    if max_t == 0:        return "— нет данных"
    if total < min_t:     return "❌ Норма не выполнена"
    elif total == min_t:  return "✅ Выполнен минимум"
    elif total < max_t:   return "👍 Выполнено средне"
    elif total == max_t:  return "🏆 Выполнен максимум"
    else:                 return "🚀 Выше максимума"

# ─── ГЕНЕРАЦИЯ PDF ────────────────────────────────────────────────────────────

EMOJI_MAP = {"➊":"NotoEmoji","🕵️":"NotoEmoji","📍":"NotoEmoji","📅":"NotoEmoji",
             "👨":"NotoEmoji","🏢":"NotoEmoji","🎓":"NotoEmoji","✍️":"NotoEmoji",
             "✔️":"NotoEmoji","📊":"NotoEmoji","📈":"NotoEmoji","📢":"NotoEmoji"}

def _tag_emojis(text):
    for e, f in EMOJI_MAP.items():
        text = text.replace(e, f"<font name='{f}'>{e}</font>")
    return text

def generate_pdf(text, filename):
    doc = SimpleDocTemplate(filename, pagesize=A4)
    styles = getSampleStyleSheet()
    s = styles["Normal"]
    s.fontName = "NotoSans"; s.fontSize = 10; s.leading = 14
    doc.build([Paragraph(_tag_emojis(text).replace("\n", "<br/>"), s)])
    auto_delete(filename)

def _calculate_plan(visits, period_name, start_dt, end_dt):
    if hasattr(start_dt, 'tzinfo') and start_dt.tzinfo: start_dt = start_dt.replace(tzinfo=None)
    if hasattr(end_dt,   'tzinfo') and end_dt.tzinfo:   end_dt   = end_dt.replace(tzinfo=None)
    if period_name == "день":     days = 1
    elif period_name == "неделю": days = 5
    else: days = _working_days_completed(start_dt, end_dt)
    doc   = len([v for v in visits if v[3] == "🩺 Врач"])
    ph    = len([v for v in visits if v[3] in ["💊 Аптека", "🚚 Дистрибьютор"]])
    ab    = len([v for v in visits if v[3] == "Не вышел"])
    wd    = max(days - ab, 1) if period_name != "день" else 1
    mn_d, mx_d = 8*wd, 12*wd
    mn_p, mx_p = 6*wd, 10*wd
    def p(v, t): return v/t*100 if t else 0
    return (
        f"\n📊 Выполнение за {period_name}:\n"
        f"👨 Врачи: Min {doc}/{mn_d} — {p(doc,mn_d):.1f}%, Max {doc}/{mx_d} — {p(doc,mx_d):.1f}%\n"
        f"🏢 Аптеки/Оптом: Min {ph}/{mn_p} — {p(ph,mn_p):.1f}%, Max {ph}/{mx_p} — {p(ph,mx_p):.1f}%\n"
        f"📈 Итого: Min {doc+ph}/{mn_d+mn_p} — {p(doc+ph,mn_d+mn_p):.1f}%, "
        f"Max {doc+ph}/{mx_d+mx_p} — {p(doc+ph,mx_d+mx_p):.1f}%\n"
    )

def build_report_text(user_info, visits, period_name, start_dt, end_dt):
    header = f"Отчёт: {user_info[1]}  {user_info[3]}  {user_info[2]}\n" if user_info else "Отчёт\n"
    if hasattr(start_dt, 'tzinfo') and start_dt.tzinfo: start_dt = start_dt.replace(tzinfo=None)
    if hasattr(end_dt,   'tzinfo') and end_dt.tzinfo:   end_dt   = end_dt.replace(tzinfo=None)
    header += f"Период: {start_dt.strftime('%d.%m.%Y')} — {end_dt.strftime('%d.%m.%Y')}\n"
    body = ""
    for i, v in enumerate(visits, 1):
        if v[3] == "Не вышел":
            body += f"===\n{v[4]} – 📢 Не вышел. Причина: {v[9]}\n===\n"
        else:
            loc = f"https://www.google.com/maps/place/{v[10]},{v[11]}" if v[10] and v[11] else "Нет данных"
            body += (f"\n➊ Визит №{i}\n📍 {v[3]}\n📅 {v[4]}\n👨 {v[5]}\n"
                     f"🏢 {v[6]}\n🎓 {v[7]}\n✍️ {v[8]}\n✔️ {v[9]}\n📌 {loc}\n===\n")
    body += _calculate_plan(visits, period_name, start_dt, end_dt)
    return header + body

# ─── ГЕНЕРАЦИЯ EXCEL (сводная) ────────────────────────────────────────────────

def generate_excel(users, by_user, period_name, start_dt, end_dt, filename):
    wb  = openpyxl.Workbook()
    hf  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hfl = PatternFill("solid", start_color="1F4E79")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thn = Side(style="thin", color="AAAAAA")
    brd = Border(left=thn, right=thn, top=thn, bottom=thn)

    ws = wb.active; ws.title = "Все визиты"
    headers = ["№","Дата/время","Сотрудник","Регион","Категория","ФИО клиента",
               "Организация","Специальность","Тема","Результат","Широта","Долгота"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font=hf; cell.fill=hfl; cell.alignment=ctr; cell.border=brd

    user_map = {u[0]: u for u in users}
    rn = 2
    for uid, visits in by_user.items():
        u = user_map.get(uid, (uid,"—","—","—"))
        for i, v in enumerate(visits, 1):
            ws.append([i,v[4],u[1],u[2],v[3],v[5],v[6],v[7],v[8],v[9],v[10],v[11]])
            for cell in ws[rn]:
                cell.font=Font(name="Arial",size=10); cell.alignment=lft; cell.border=brd
            rn += 1

    for i, w in enumerate([4,18,25,15,18,25,25,20,30,30,12,12],1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Сводка
    ws2 = wb.create_sheet("Сводка")
    ws2.append(["Сотрудник","Регион","Визитов","Врачи","Аптеки","Не вышел","% (min)","% (max)"])
    for cell in ws2[1]:
        cell.font=hf; cell.fill=hfl; cell.alignment=ctr; cell.border=brd

    if period_name == "день":     days = 1
    elif period_name == "неделю": days = 5
    else:
        if hasattr(start_dt,'tzinfo') and start_dt.tzinfo: start_dt=start_dt.replace(tzinfo=None)
        if hasattr(end_dt,'tzinfo')   and end_dt.tzinfo:   end_dt=end_dt.replace(tzinfo=None)
        days = _working_days_completed(start_dt, end_dt)

    for ri, (uid, visits) in enumerate(by_user.items(), 2):
        u   = user_map.get(uid, (uid,"—","—","—"))
        doc = len([v for v in visits if v[3]=="🩺 Врач"])
        ph  = len([v for v in visits if v[3] in ["💊 Аптека","🚚 Дистрибьютор"]])
        ab  = len([v for v in visits if v[3]=="Не вышел"])
        tot = doc + ph
        wd  = max(days-ab,1)
        pm  = round(tot/max((8+6)*wd,1)*100,1)
        px  = round(tot/max((12+10)*wd,1)*100,1)
        ws2.append([u[1],u[2],len(visits),doc,ph,ab,f"{pm}%",f"{px}%"])
        fill = PatternFill("solid", start_color="E2EFDA" if pm>=100 else "FCE4D6" if pm<60 else "FFEB9C")
        for cell in ws2[ri]:
            cell.font=Font(name="Arial",size=10); cell.alignment=ctr; cell.border=brd; cell.fill=fill

    for i,w in enumerate([25,15,14,10,14,10,16,16],1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"
    wb.save(filename); auto_delete(filename)

# ─── ОТЧЁТ ПО КОМАНДЕ (Excel) ─────────────────────────────────────────────────

def generate_team_report_excel(users_with_status, visits_db, users_db, start_dt, end_dt, period_name, filename):
    if hasattr(start_dt,'tzinfo') and start_dt.tzinfo: start_dt=start_dt.replace(tzinfo=None)
    if hasattr(end_dt,'tzinfo')   and end_dt.tzinfo:   end_dt=end_dt.replace(tzinfo=None)
    today = datetime.now().replace(hour=0,minute=0,second=0,microsecond=0)
    if end_dt >= today: end_dt = today - timedelta(seconds=1)

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Отчёт по команде"
    hf  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hfl = PatternFill("solid", start_color="1F4E79")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    lft = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thn = Side(style="thin", color="AAAAAA")
    brd = Border(left=thn, right=thn, top=thn, bottom=thn)

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Отчёт по команде | {start_dt.strftime('%d.%m.%Y')} — {end_dt.strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = ctr
    ws["A1"].fill = PatternFill("solid", start_color="DEEAF1")
    ws.row_dimensions[1].height = 28

    headers = ["№", "ФИО", "Регион", "Начало работы", "Вердикт %", "День рождения", "Телефон", "Статус", "Дней работает"]
    ws.append(headers)
    for cell in ws[2]:
        cell.font=hf; cell.fill=hfl; cell.alignment=ctr; cell.border=brd
    ws.row_dimensions[2].height = 36

    wd = _working_days_completed(start_dt, end_dt)

    for idx, u in enumerate(users_with_status, 1):
        # u = (telegram_id, full_name, region, phone_number, is_active)
        uid      = u[0]
        # Получаем дату рождения из users.db
        with sqlite3.connect(users_db) as conn:
            cur = conn.cursor()
            cur.execute("SELECT birth_date FROM users WHERE telegram_id=?", (uid,))
            row = cur.fetchone()
        birth_date = row[0] if row else "—"

        # Визиты за период
        with sqlite3.connect(visits_db) as conn:
            cur = conn.cursor()
            cur.execute("SELECT * FROM visits WHERE telegram_id=?", (uid,))
            all_v = cur.fetchall()
        visits = []
        for v in all_v:
            try:
                dt = datetime.strptime(v[4], "%d.%m.%Y %H:%M")
                if start_dt <= dt <= end_dt: visits.append(v)
            except Exception: continue

        absent  = len([v for v in visits if v[3] == "Не вышел"])
        wd_u    = max(wd - absent, 1) if wd > 0 else 1
        doc     = len([v for v in visits if v[3] == "🩺 Врач"])
        ph      = len([v for v in visits if v[3] in ["💊 Аптека","🚚 Дистрибьютор"]])
        total   = doc + ph
        min_t   = (8+6)*wd_u
        max_t   = (12+10)*wd_u
        pct_str = f"{_pct(total, min_t)}%"
        verdict = _verdict_short(total, min_t, max_t)
        verd_str = f"{pct_str} — {verdict}"

        first_date = get_first_visit_date(visits_db, uid)
        days_worked= get_days_worked(visits_db, uid)
        status     = "✅ Работает" if u[4] else "🚫 Уволен"

        ws.append([idx, u[1], u[2], first_date or "—", verd_str, birth_date, u[3], status, days_worked])

        # Цвет строки
        if not u[4]:
            row_fill = PatternFill("solid", start_color="FFE0E0")
        elif "не выполнена" in verdict:
            row_fill = PatternFill("solid", start_color="FF7676")
        elif "минимум" in verdict:
            row_fill = PatternFill("solid", start_color="FFEB9C")
        elif "средне" in verdict:
            row_fill = PatternFill("solid", start_color="FCE4D6")
        else:
            row_fill = PatternFill("solid", start_color="E2EFDA")

        for ci, cell in enumerate(ws[idx+2], 1):
            cell.font = Font(name="Arial", size=10)
            cell.fill = row_fill
            cell.border = brd
            cell.alignment = lft if ci in (2, 3, 5, 7) else ctr

    for i, w in enumerate([4, 28, 16, 16, 36, 14, 16, 14, 14], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    wb.save(filename); auto_delete(filename)

# ─── КЛАВИАТУРЫ ───────────────────────────────────────────────────────────────

def kb_main():
    return ReplyKeyboardMarkup([
        ["👤 Отчёт по сотруднику"],
        ["📊 Сводный отчёт (все)"],
        ["📥 Выгрузка в Excel"],
        ["📅 Тур-план сотрудника"],
        ["👥 Список сотрудников"],
        ["📈 Аналитика"],
        ["⚙️ Управление"],
    ], resize_keyboard=True)

def kb_period():
    return ReplyKeyboardMarkup([
        ["🗓 За текущий день"],
        ["📆 За текущую неделю"],
        ["📊 За текущий месяц"],
        ["📅 Свободный выбор дат"],
        ["🔙 Назад"],
    ], resize_keyboard=True, one_time_keyboard=True)

def kb_employees_active():
    users = get_active_users(USERS_DB)
    rows  = [[f"{u[1]} ({u[2]})"] for u in users]
    rows.append(["🔙 Назад"])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True, one_time_keyboard=True), users

def kb_employees_all():
    users = get_all_users_with_status(USERS_DB)
    rows  = [[f"{'✅' if u[4] else '🚫'} {u[1]} ({u[2]})"] for u in users]
    rows.append(["🔙 Назад"])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True, one_time_keyboard=True), users

def kb_mgmt():
    return ReplyKeyboardMarkup([
        ["👥 Команда", "📋 Отчёт по команде"],
        ["🔙 Назад"],
    ], resize_keyboard=True)

def kb_team_action():
    return ReplyKeyboardMarkup([
        ["🚫 Отключить", "✅ Восстановить"],
        ["🔙 Назад"],
    ], resize_keyboard=True, one_time_keyboard=True)

# ─── ХЕНДЛЕРЫ ─────────────────────────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update.effective_user.id):
        await update.message.reply_text("⛔ Нет доступа.")
        return ConversationHandler.END
    await update.message.reply_text("👑 Добро пожаловать в Админ-панель!", reply_markup=kb_main())
    return MAIN_MENU

async def main_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update.effective_user.id): return ConversationHandler.END
    text = update.message.text

    if text == "👥 Список сотрудников":
        users = get_active_users(USERS_DB)
        if not users:
            await update.message.reply_text("Нет сотрудников.", reply_markup=kb_main()); return MAIN_MENU
        msg = "👥 *Список сотрудников:*\n\n"
        for i, u in enumerate(users, 1):
            msg += f"{i}. {u[1]}\n   📍 {u[2]}  |  📞 {u[3]}\n"
        await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=kb_main())
        return MAIN_MENU

    elif text == "👤 Отчёт по сотруднику":
        kb, users = kb_employees_active()
        if not users:
            await update.message.reply_text("Нет сотрудников.", reply_markup=kb_main()); return MAIN_MENU
        context.user_data["users_list"] = users
        await update.message.reply_text("Выберите сотрудника:", reply_markup=kb)
        return CHOOSE_EMPLOYEE

    elif text == "📊 Сводный отчёт (все)":
        await update.message.reply_text("Выберите период:", reply_markup=kb_period())
        return SUMMARY_PERIOD

    elif text == "📥 Выгрузка в Excel":
        await update.message.reply_text("Выберите период:", reply_markup=kb_period())
        return EXCEL_PERIOD

    elif text == "📅 Тур-план сотрудника":
        kb, users = kb_employees_active()
        if not users:
            await update.message.reply_text("Нет сотрудников.", reply_markup=kb_main()); return MAIN_MENU
        context.user_data["users_list"] = users
        await update.message.reply_text("Выберите сотрудника:", reply_markup=kb)
        return TOURPLAN_EMPLOYEE

    elif text == "📈 Аналитика":
        await update.message.reply_text("Выберите период:", reply_markup=kb_period())
        return ANALYTICS_PERIOD

    elif text == "⚙️ Управление":
        await update.message.reply_text("⚙️ Управление:", reply_markup=kb_mgmt())
        return MGMT_MENU

    return MAIN_MENU

# ── Отчёт по сотруднику ───────────────────────────────────────────────────────

async def choose_employee(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Главное меню:", reply_markup=kb_main()); return MAIN_MENU
    users  = context.user_data.get("users_list", [])
    chosen = next((u for u in users if f"{u[1]} ({u[2]})" == text), None)
    if not chosen:
        await update.message.reply_text("Не найден. Попробуйте снова."); return CHOOSE_EMPLOYEE
    context.user_data["chosen_user"] = chosen
    await update.message.reply_text(f"✅ {chosen[1]}\nВыберите период:", reply_markup=kb_period())
    return CHOOSE_PERIOD

async def choose_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Главное меню:", reply_markup=kb_main()); return MAIN_MENU
    if text == "📅 Свободный выбор дат":
        await update.message.reply_text("Введите дату начала (ДД.ММ.ГГГГ):"); return CHOOSE_START
    s, e, pn = period_bounds(text)
    await _send_employee_report(update, context, s, e, pn); return MAIN_MENU

async def choose_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["custom_start"] = datetime.strptime(update.message.text, "%d.%m.%Y")
        await update.message.reply_text("Введите дату окончания (ДД.ММ.ГГГГ):"); return CHOOSE_END
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ"); return CHOOSE_START

async def choose_end(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        e = datetime.strptime(update.message.text, "%d.%m.%Y").replace(hour=23,minute=59,second=59)
        s = context.user_data["custom_start"]
        if e < s:
            await update.message.reply_text("❌ Дата окончания раньше начала."); return CHOOSE_END
        await _send_employee_report(update, context, s, e, "период"); return MAIN_MENU
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ"); return CHOOSE_END

async def _send_employee_report(update, context, start_dt, end_dt, period_name):
    user   = context.user_data["chosen_user"]
    visits = get_visits_for_user(user[0], start_dt, end_dt)
    await update.message.reply_text(f"⏳ Формирую отчёт по {user[1]}...")
    text = build_report_text(user, visits, period_name, start_dt, end_dt)
    fn   = os.path.join(REPORTS_DIR, f"report_{user[0]}_{datetime.now().strftime('%d%m%Y%H%M%S')}.pdf")
    generate_pdf(text, fn)
    with open(fn, "rb") as f:
        await update.message.reply_document(f, caption=f"📄 {user[1]} за {period_name}")
    await update.message.reply_text("Главное меню:", reply_markup=kb_main())

# ── Сводный отчёт ─────────────────────────────────────────────────────────────

async def summary_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Главное меню:", reply_markup=kb_main()); return MAIN_MENU
    if text == "📅 Свободный выбор дат":
        await update.message.reply_text("Введите дату начала (ДД.ММ.ГГГГ):"); return SUMMARY_START
    s, e, pn = period_bounds(text)
    await _send_summary(update, context, s, e, pn); return MAIN_MENU

async def summary_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["sum_start"] = datetime.strptime(update.message.text, "%d.%m.%Y")
        await update.message.reply_text("Введите дату окончания (ДД.ММ.ГГГГ):"); return SUMMARY_END
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ"); return SUMMARY_START

async def summary_end(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        e = datetime.strptime(update.message.text, "%d.%m.%Y").replace(hour=23,minute=59,second=59)
        s = context.user_data["sum_start"]
        if e < s:
            await update.message.reply_text("❌ Дата окончания раньше начала."); return SUMMARY_END
        await _send_summary(update, context, s, e, "период"); return MAIN_MENU
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ"); return SUMMARY_END

async def _send_summary(update, context, start_dt, end_dt, period_name):
    users  = get_active_users(USERS_DB)
    visits = get_all_visits_filtered(start_dt, end_dt)
    await update.message.reply_text("⏳ Формирую сводный отчёт...")
    by_user = {}
    for v in visits: by_user.setdefault(v[1], []).append(v)
    user_map = {u[0]: u for u in users}
    full = f"📊 СВОДНЫЙ ОТЧЁТ за {period_name}\n" + "="*40 + "\n"
    for uid, uvs in by_user.items():
        uinfo = user_map.get(uid, (uid,"—","—","—"))
        full += f"\n👤 {uinfo[1]} ({uinfo[2]})\n"
        full += build_report_text(uinfo, uvs, period_name, start_dt, end_dt)
        full += "\n" + "="*40 + "\n"
    fn = os.path.join(REPORTS_DIR, f"summary_{datetime.now().strftime('%d%m%Y%H%M%S')}.pdf")
    generate_pdf(full, fn)
    with open(fn, "rb") as f:
        await update.message.reply_document(f, caption=f"📊 Сводный за {period_name}")
    await update.message.reply_text("Главное меню:", reply_markup=kb_main())

# ── Excel выгрузка ────────────────────────────────────────────────────────────

async def excel_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Главное меню:", reply_markup=kb_main()); return MAIN_MENU
    if text == "📅 Свободный выбор дат":
        await update.message.reply_text("Введите дату начала (ДД.ММ.ГГГГ):"); return EXCEL_START
    s, e, pn = period_bounds(text)
    await _send_excel(update, context, s, e, pn); return MAIN_MENU

async def excel_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["excel_start"] = datetime.strptime(update.message.text, "%d.%m.%Y")
        await update.message.reply_text("Введите дату окончания (ДД.ММ.ГГГГ):"); return EXCEL_END
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ"); return EXCEL_START

async def excel_end(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        e = datetime.strptime(update.message.text, "%d.%m.%Y").replace(hour=23,minute=59,second=59)
        s = context.user_data["excel_start"]
        if e < s:
            await update.message.reply_text("❌ Дата окончания раньше начала."); return EXCEL_END
        await _send_excel(update, context, s, e, "период"); return MAIN_MENU
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ"); return EXCEL_END

async def _send_excel(update, context, start_dt, end_dt, period_name):
    users  = get_active_users(USERS_DB)
    visits = get_all_visits_filtered(start_dt, end_dt)
    await update.message.reply_text("⏳ Формирую Excel...")
    by_user = {}
    for v in visits: by_user.setdefault(v[1], []).append(v)
    fn = os.path.join(REPORTS_DIR, f"excel_{datetime.now().strftime('%d%m%Y%H%M%S')}.xlsx")
    generate_excel(users, by_user, period_name, start_dt, end_dt, fn)
    with open(fn, "rb") as f:
        await update.message.reply_document(f, caption=f"📥 Excel за {period_name}")
    await update.message.reply_text("Главное меню:", reply_markup=kb_main())

# ── Тур-план ─────────────────────────────────────────────────────────────────

async def tourplan_employee(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Главное меню:", reply_markup=kb_main()); return MAIN_MENU
    users  = context.user_data.get("users_list", [])
    chosen = next((u for u in users if f"{u[1]} ({u[2]})" == text), None)
    if not chosen:
        await update.message.reply_text("Не найден."); return TOURPLAN_EMPLOYEE
    plan = get_tourplan(chosen[0])
    if not plan:
        await update.message.reply_text(f"У {chosen[1]} нет тур-плана.", reply_markup=kb_main()); return MAIN_MENU
    msg = f"📅 *Тур-план: {chosen[1]}*\n\n"
    for r in plan: msg += f"📆 {r[1]} | 📍 {r[2]}\n🏢 {r[3]}\n🎯 {r[4]}\n\n"
    await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=kb_main())
    return MAIN_MENU

# ── Аналитика ────────────────────────────────────────────────────────────────

async def analytics_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Главное меню:", reply_markup=kb_main()); return MAIN_MENU
    if text == "📅 Свободный выбор дат":
        await update.message.reply_text("Введите дату начала (ДД.ММ.ГГГГ):"); return ANALYTICS_START
    s, e, pn = period_bounds(text)
    await _send_analytics(update, context, s, e, pn); return MAIN_MENU

async def analytics_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["analytics_start"] = datetime.strptime(update.message.text, "%d.%m.%Y")
        await update.message.reply_text("Введите дату окончания (ДД.ММ.ГГГГ):"); return ANALYTICS_END
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ"); return ANALYTICS_START

async def analytics_end(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        e = datetime.strptime(update.message.text, "%d.%m.%Y").replace(hour=23,minute=59,second=59)
        s = context.user_data["analytics_start"]
        if e < s:
            await update.message.reply_text("❌ Дата окончания раньше начала."); return ANALYTICS_END
        await _send_analytics(update, context, s, e, "период"); return MAIN_MENU
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ"); return ANALYTICS_END

async def _send_analytics(update, context, start_dt, end_dt, period_name):
    await update.message.reply_text("⏳ Формирую аналитику...")
    fn = os.path.join(REPORTS_DIR, f"analytics_{datetime.now().strftime('%d%m%Y%H%M%S')}.xlsx")
    generate_analytics_excel(VISITS_DB, USERS_DB, start_dt, end_dt, period_name, fn)
    with open(fn, "rb") as f:
        await update.message.reply_document(f, caption=f"📈 Аналитика за {period_name}")
    auto_delete(fn)
    await update.message.reply_text("Главное меню:", reply_markup=kb_main())

# ── Управление ────────────────────────────────────────────────────────────────

async def mgmt_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Главное меню:", reply_markup=kb_main()); return MAIN_MENU

    elif text == "👥 Команда":
        kb, users = kb_employees_all()
        context.user_data["all_users_with_status"] = users
        await update.message.reply_text(
            "Выберите сотрудника (✅ — работает, 🚫 — уволен):", reply_markup=kb)
        return TEAM_CHOOSE_EMPLOYEE

    elif text == "📋 Отчёт по команде":
        await update.message.reply_text("Выберите период:", reply_markup=kb_period())
        return TEAM_REPORT_PERIOD

    return MGMT_MENU

async def team_choose_employee(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Управление:", reply_markup=kb_mgmt()); return MGMT_MENU

    users   = context.user_data.get("all_users_with_status", [])
    # Убираем статус-префикс из имени
    clean   = text.replace("✅ ", "").replace("🚫 ", "")
    chosen  = next((u for u in users if f"{u[1]} ({u[2]})" == clean), None)
    if not chosen:
        await update.message.reply_text("Не найден. Попробуйте снова."); return TEAM_CHOOSE_EMPLOYEE

    context.user_data["team_chosen"] = chosen
    status = "✅ Работает" if chosen[4] else "🚫 Уволен"
    await update.message.reply_text(
        f"👤 {chosen[1]}\n📍 {chosen[2]}\nСтатус: {status}\n\nВыберите действие:",
        reply_markup=kb_team_action()
    )
    return TEAM_ACTION

async def team_action(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text   = update.message.text
    chosen = context.user_data.get("team_chosen")
    if not chosen:
        await update.message.reply_text("Ошибка.", reply_markup=kb_main()); return MAIN_MENU

    if text == "🔙 Назад":
        await update.message.reply_text("Управление:", reply_markup=kb_mgmt()); return MGMT_MENU

    elif text == "🚫 Отключить":
        deactivate_employee(chosen[0])
        await update.message.reply_text(
            f"🚫 {chosen[1]} помечен как уволенный.\nОн больше не появится в аналитике и отчётах.",
            reply_markup=kb_mgmt()
        )
    elif text == "✅ Восстановить":
        activate_employee(chosen[0])
        await update.message.reply_text(
            f"✅ {chosen[1]} восстановлен и снова активен.",
            reply_markup=kb_mgmt()
        )

    return MGMT_MENU

# ── Отчёт по команде ──────────────────────────────────────────────────────────

async def team_report_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Управление:", reply_markup=kb_mgmt()); return MGMT_MENU
    if text == "📅 Свободный выбор дат":
        await update.message.reply_text("Введите дату начала (ДД.ММ.ГГГГ):"); return TEAM_REPORT_START
    s, e, pn = period_bounds(text)
    await _send_team_report(update, context, s, e, pn); return MGMT_MENU

async def team_report_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["tr_start"] = datetime.strptime(update.message.text, "%d.%m.%Y")
        await update.message.reply_text("Введите дату окончания (ДД.ММ.ГГГГ):"); return TEAM_REPORT_END
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ"); return TEAM_REPORT_START

async def team_report_end(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        e = datetime.strptime(update.message.text, "%d.%m.%Y").replace(hour=23,minute=59,second=59)
        s = context.user_data["tr_start"]
        if e < s:
            await update.message.reply_text("❌ Дата окончания раньше начала."); return TEAM_REPORT_END
        await _send_team_report(update, context, s, e, "период"); return MGMT_MENU
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ"); return TEAM_REPORT_END

async def _send_team_report(update, context, start_dt, end_dt, period_name):
    users = get_all_users_with_status(USERS_DB)
    await update.message.reply_text("⏳ Формирую отчёт по команде...")
    fn = os.path.join(REPORTS_DIR, f"team_{datetime.now().strftime('%d%m%Y%H%M%S')}.xlsx")
    generate_team_report_excel(users, VISITS_DB, USERS_DB, start_dt, end_dt, period_name, fn)
    with open(fn, "rb") as f:
        await update.message.reply_document(f, caption=f"📋 Отчёт по команде за {period_name}")
    auto_delete(fn)
    await update.message.reply_text("Управление:", reply_markup=kb_mgmt())

# ─── ЗАПУСК ───────────────────────────────────────────────────────────────────

def main():
    app = Application.builder().token(ADMIN_BOT_TOKEN).read_timeout(120).write_timeout(120).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            MAIN_MENU:            [MessageHandler(filters.TEXT & ~filters.COMMAND, main_menu_handler)],
            CHOOSE_EMPLOYEE:      [MessageHandler(filters.TEXT & ~filters.COMMAND, choose_employee)],
            CHOOSE_PERIOD:        [MessageHandler(filters.TEXT & ~filters.COMMAND, choose_period)],
            CHOOSE_START:         [MessageHandler(filters.TEXT & ~filters.COMMAND, choose_start)],
            CHOOSE_END:           [MessageHandler(filters.TEXT & ~filters.COMMAND, choose_end)],
            SUMMARY_PERIOD:       [MessageHandler(filters.TEXT & ~filters.COMMAND, summary_period)],
            SUMMARY_START:        [MessageHandler(filters.TEXT & ~filters.COMMAND, summary_start)],
            SUMMARY_END:          [MessageHandler(filters.TEXT & ~filters.COMMAND, summary_end)],
            EXCEL_PERIOD:         [MessageHandler(filters.TEXT & ~filters.COMMAND, excel_period)],
            EXCEL_START:          [MessageHandler(filters.TEXT & ~filters.COMMAND, excel_start)],
            EXCEL_END:            [MessageHandler(filters.TEXT & ~filters.COMMAND, excel_end)],
            TOURPLAN_EMPLOYEE:    [MessageHandler(filters.TEXT & ~filters.COMMAND, tourplan_employee)],
            ANALYTICS_PERIOD:     [MessageHandler(filters.TEXT & ~filters.COMMAND, analytics_period)],
            ANALYTICS_START:      [MessageHandler(filters.TEXT & ~filters.COMMAND, analytics_start)],
            ANALYTICS_END:        [MessageHandler(filters.TEXT & ~filters.COMMAND, analytics_end)],
            MGMT_MENU:            [MessageHandler(filters.TEXT & ~filters.COMMAND, mgmt_menu)],
            TEAM_CHOOSE_EMPLOYEE: [MessageHandler(filters.TEXT & ~filters.COMMAND, team_choose_employee)],
            TEAM_ACTION:          [MessageHandler(filters.TEXT & ~filters.COMMAND, team_action)],
            TEAM_REPORT_PERIOD:   [MessageHandler(filters.TEXT & ~filters.COMMAND, team_report_period)],
            TEAM_REPORT_START:    [MessageHandler(filters.TEXT & ~filters.COMMAND, team_report_start)],
            TEAM_REPORT_END:      [MessageHandler(filters.TEXT & ~filters.COMMAND, team_report_end)],
        },
        fallbacks=[CommandHandler("start", start)],
    )

    app.add_handler(conv)
    print("✅ Админ-бот запущен")
    app.run_polling()

if __name__ == "__main__":
    main()
