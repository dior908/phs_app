import os
import sqlite3
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import pytz
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import threading

from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application, CommandHandler, MessageHandler, filters,
    ConversationHandler, ContextTypes
)

# ─── КОНФИГ ───────────────────────────────────────────────────────────────────
ADMIN_BOT_TOKEN = "8341249968:AAG55xcFQdl-54d30fvE9OLVX-bhI8FiUsU"  # ← замени на новый токен

# ID пользователей, которым разрешён доступ (добавляй HR/директора сюда)
ALLOWED_IDS = [6669377232]

# Путь к папке основного бота (где лежат databases/)
MAIN_BOT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)))

# Пути к базам данных основного бота
USERS_DB  = os.path.join(MAIN_BOT_DIR, "databases", "users.db")
VISITS_DB = os.path.join(MAIN_BOT_DIR, "databases", "visits.db")
TOURPLAN_DB = os.path.join(MAIN_BOT_DIR, "databases", "tourplan.db")

# Папка для временных файлов отчётов
REPORTS_DIR = os.path.join(MAIN_BOT_DIR, "admin_reports")
os.makedirs(REPORTS_DIR, exist_ok=True)

# Шрифты (берём из основного бота)
FONTS_DIR = os.path.join(MAIN_BOT_DIR, "fonts")
NOTO_SANS  = os.path.join(FONTS_DIR, "NotoSans-Regular.ttf")
NOTO_EMOJI = os.path.join(FONTS_DIR, "NotoEmoji-Regular.ttf")

if os.path.exists(NOTO_SANS) and os.path.exists(NOTO_EMOJI):
    pdfmetrics.registerFont(TTFont("NotoSans",  NOTO_SANS))
    pdfmetrics.registerFont(TTFont("NotoEmoji", NOTO_EMOJI))

# ─── СОСТОЯНИЯ ConversationHandler ────────────────────────────────────────────
(
    MAIN_MENU,
    CHOOSE_EMPLOYEE, CHOOSE_PERIOD, CHOOSE_START, CHOOSE_END,
    SUMMARY_PERIOD, SUMMARY_START, SUMMARY_END,
    EXCEL_PERIOD, EXCEL_START, EXCEL_END,
    TOURPLAN_EMPLOYEE,
) = range(12)

TZ = pytz.timezone("Asia/Tashkent")

# ─── ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ──────────────────────────────────────────────────

def is_allowed(user_id: int) -> bool:
    return user_id in ALLOWED_IDS


def get_all_users():
    with sqlite3.connect(USERS_DB) as conn:
        cur = conn.cursor()
        cur.execute("SELECT telegram_id, full_name, region, phone_number FROM users ORDER BY full_name")
        return cur.fetchall()


def get_visits(telegram_id, start_dt: datetime, end_dt: datetime):
    start_s = start_dt.strftime("%d.%m.%Y %H:%M")
    end_s   = end_dt.strftime("%d.%m.%Y %H:%M")
    with sqlite3.connect(VISITS_DB) as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT * FROM visits WHERE telegram_id = ? AND date_time BETWEEN ? AND ?",
            (telegram_id, start_s, end_s)
        )
        return cur.fetchall()


def get_all_visits(start_dt: datetime, end_dt: datetime):
    start_s = start_dt.strftime("%d.%m.%Y %H:%M")
    end_s   = end_dt.strftime("%d.%m.%Y %H:%M")
    with sqlite3.connect(VISITS_DB) as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT * FROM visits WHERE date_time BETWEEN ? AND ? ORDER BY telegram_id, date_time",
            (start_s, end_s)
        )
        return cur.fetchall()


def get_tourplan(telegram_id):
    with sqlite3.connect(TOURPLAN_DB) as conn:
        cur = conn.cursor()
        cur.execute("SELECT * FROM tourplan WHERE telegram_id = ? ORDER BY date", (telegram_id,))
        return cur.fetchall()


def period_bounds(period: str):
    """Возвращает (start_dt, end_dt, period_name) для стандартных периодов."""
    today = datetime.now(TZ)
    if period == "🗓 За текущий день":
        s = today.replace(hour=0, minute=1, second=0, microsecond=0)
        return s, today, "день"
    elif period == "📆 За текущую неделю":
        s = (today - timedelta(days=today.weekday())).replace(hour=0, minute=1, second=0, microsecond=0)
        return s, today, "неделю"
    elif period == "📊 За текущий месяц":
        s = today.replace(day=1, hour=0, minute=1, second=0, microsecond=0)
        return s, today, "месяц"
    return None, None, None


def calculate_plan_percentage(visits, period_name, start_dt, end_dt):
    if period_name == "день":
        days = 1
    elif period_name == "неделю":
        days = 5
    else:
        days = sum(
            1 for d in range((end_dt - start_dt).days + 1)
            if (start_dt + timedelta(days=d)).weekday() < 5
        )

    doctor_visits   = len([v for v in visits if v[3] == "🩺 Врач"])
    pharm_visits    = len([v for v in visits if v[3] in ["💊 Аптека", "🚚 Дистрибьютор"]])
    absences        = len([v for v in visits if v[3] == "Не вышел"])
    working_days    = max(days - absences, 1) if period_name != "день" else 1

    min_doc, max_doc   = 8 * working_days, 12 * working_days
    min_ph,  max_ph    = 6 * working_days, 10 * working_days

    def pct(val, total): return val / total * 100 if total else 0

    return (
        f"\n📊 Выполнение плана за {period_name}:\n"
        f"👨 Врачи:          Min {doctor_visits}/{min_doc} — {pct(doctor_visits, min_doc):.1f}%,"
        f" Max {doctor_visits}/{max_doc} — {pct(doctor_visits, max_doc):.1f}%\n"
        f"🏢 Аптеки/Оптом:   Min {pharm_visits}/{min_ph} — {pct(pharm_visits, min_ph):.1f}%,"
        f" Max {pharm_visits}/{max_ph} — {pct(pharm_visits, max_ph):.1f}%\n"
        f"📈 Итого:           Min {doctor_visits+pharm_visits}/{min_doc+min_ph} —"
        f" {pct(doctor_visits+pharm_visits, min_doc+min_ph):.1f}%,"
        f" Max {doctor_visits+pharm_visits}/{max_doc+max_ph} —"
        f" {pct(doctor_visits+pharm_visits, max_doc+max_ph):.1f}%\n"
    )


def auto_delete(path, seconds=3600):
    threading.Timer(seconds, lambda: os.path.exists(path) and os.remove(path)).start()


# ─── ГЕНЕРАЦИЯ PDF ────────────────────────────────────────────────────────────

EMOJI_MAP = {
    "➊":"NotoEmoji","🕵️":"NotoEmoji","📍":"NotoEmoji","📅":"NotoEmoji",
    "👨":"NotoEmoji","🏢":"NotoEmoji","🎓":"NotoEmoji","✍️":"NotoEmoji",
    "✔️":"NotoEmoji","📊":"NotoEmoji","📈":"NotoEmoji","📢":"NotoEmoji",
}

def _tag_emojis(text):
    for emoji, font in EMOJI_MAP.items():
        text = text.replace(emoji, f"<font name='{font}'>{emoji}</font>")
    return text


def generate_pdf(report_text: str, filename: str):
    doc = SimpleDocTemplate(filename, pagesize=A4)
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "NotoSans"
    style.fontSize = 10
    style.leading  = 14
    story = [Paragraph(_tag_emojis(report_text).replace("\n", "<br/>"), style)]
    doc.build(story)
    auto_delete(filename)


def build_report_text(user_info, visits, period_name, start_dt, end_dt):
    if user_info:
        header = f"Отчёт: {user_info[1]}  {user_info[3]}  {user_info[2]}\n"
    else:
        header = "Отчёт (пользователь не найден)\n"
    header += f"Период: {start_dt.strftime('%d.%m.%Y')} — {end_dt.strftime('%d.%m.%Y')}\n"

    body = ""
    for i, v in enumerate(visits, 1):
        if v[3] == "Не вышел":
            body += f"===================================\n{v[4]} – 📢 Не вышел. Причина: {v[9]}\n===================================\n"
        else:
            loc = f"https://www.google.com/maps/place/{v[10]},{v[11]}" if v[10] and v[11] else "Нет данных"
            body += (
                f"\n➊ Визит №{i}\n"
                f"🕵️ ID: {v[0]}\n"
                f"📍 Категория: {v[3]}\n"
                f"📅 Дата/время: {v[4]}\n"
                f"👨 ФИО: {v[5]}\n"
                f"🏢 Организация: {v[6]}\n"
                f"🎓 Специальность: {v[7]}\n"
                f"✍️ Тема: {v[8]}\n"
                f"✔️ Результат: {v[9]}\n"
                f"📌 Локация: {loc}\n"
                f"=======================================\n"
            )

    body += calculate_plan_percentage(visits, period_name, start_dt, end_dt)
    return header + body


# ─── ГЕНЕРАЦИЯ EXCEL ──────────────────────────────────────────────────────────

def generate_excel(users, all_visits_by_user, period_name, start_dt, end_dt, filename):
    wb = openpyxl.Workbook()

    # Стили
    header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill   = PatternFill("solid", start_color="1F4E79")
    sub_fill      = PatternFill("solid", start_color="BDD7EE")
    center        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left          = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin          = Side(style="thin", color="AAAAAA")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Лист 1: Все визиты ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Все визиты"

    title_row = [
        "№", "Дата/время", "Сотрудник", "Регион",
        "Категория", "ФИО клиента", "Организация",
        "Специальность", "Тема визита", "Результат",
        "Широта", "Долгота"
    ]
    ws.append(title_row)
    for col, cell in enumerate(ws[1], 1):
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    row_num = 2
    user_map = {u[0]: u for u in users}
    for uid, visits in all_visits_by_user.items():
        uinfo = user_map.get(uid, (uid, "—", "—", "—"))
        for i, v in enumerate(visits, 1):
            ws.append([
                i, v[4], uinfo[1], uinfo[2],
                v[3], v[5], v[6], v[7], v[8], v[9],
                v[10], v[11]
            ])
            for cell in ws[row_num]:
                cell.font      = Font(name="Arial", size=10)
                cell.alignment = left
                cell.border    = border
            row_num += 1

    col_widths = [4, 18, 25, 15, 18, 25, 25, 20, 30, 30, 12, 12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ── Лист 2: Сводка по сотрудникам ───────────────────────────────────────
    ws2 = wb.create_sheet("Сводка")
    summary_headers = [
        "Сотрудник", "Регион", "Визитов всего",
        "Врачи", "Аптеки/Дистр.", "Не вышел",
        "% выполн. (min)", "% выполн. (max)"
    ]
    ws2.append(summary_headers)
    for col, cell in enumerate(ws2[1], 1):
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = border

    if period_name == "день":    days = 1
    elif period_name == "неделю": days = 5
    else:
        days = sum(
            1 for d in range((end_dt - start_dt).days + 1)
            if (start_dt + timedelta(days=d)).weekday() < 5
        )

    for row_i, (uid, visits) in enumerate(all_visits_by_user.items(), 2):
        uinfo      = user_map.get(uid, (uid, "—", "—", "—"))
        doctors    = len([v for v in visits if v[3] == "🩺 Врач"])
        pharm      = len([v for v in visits if v[3] in ["💊 Аптека", "🚚 Дистрибьютор"]])
        absent     = len([v for v in visits if v[3] == "Не вышел"])
        total      = doctors + pharm
        wd         = max(days - absent, 1) if period_name != "день" else 1
        pct_min    = round(total / max((8+6)*wd, 1) * 100, 1)
        pct_max    = round(total / max((12+10)*wd, 1) * 100, 1)

        ws2.append([uinfo[1], uinfo[2], len(visits), doctors, pharm, absent,
                    f"{pct_min}%", f"{pct_max}%"])

        fill = PatternFill("solid", start_color="E2EFDA") if pct_min >= 100 else \
               PatternFill("solid", start_color="FCE4D6") if pct_min < 60  else \
               PatternFill("solid", start_color="FFEB9C")
        for cell in ws2[row_i]:
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = center
            cell.border    = border
            cell.fill      = fill

    for i, w in enumerate([25, 15, 14, 10, 14, 10, 16, 16], 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    wb.save(filename)
    auto_delete(filename)


# ─── КЛАВИАТУРЫ ───────────────────────────────────────────────────────────────

def kb_main():
    return ReplyKeyboardMarkup([
        ["👤 Отчёт по сотруднику"],
        ["📊 Сводный отчёт (все)"],
        ["📥 Выгрузка в Excel"],
        ["📅 Тур-план сотрудника"],
        ["👥 Список сотрудников"],
    ], resize_keyboard=True)


def kb_period():
    return ReplyKeyboardMarkup([
        ["🗓 За текущий день"],
        ["📆 За текущую неделю"],
        ["📊 За текущий месяц"],
        ["📅 Свободный выбор дат"],
        ["🔙 Назад"],
    ], resize_keyboard=True, one_time_keyboard=True)


def kb_employees():
    users = get_all_users()
    rows  = [[f"{u[1]} ({u[2]})"] for u in users]
    rows.append(["🔙 Назад"])
    return ReplyKeyboardMarkup(rows, resize_keyboard=True, one_time_keyboard=True), users


# ─── ХЕНДЛЕРЫ ─────────────────────────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update.effective_user.id):
        await update.message.reply_text("⛔ У вас нет доступа к этому боту.")
        return ConversationHandler.END
    await update.message.reply_text("👑 Добро пожаловать в Админ-панель!", reply_markup=kb_main())
    return MAIN_MENU


async def main_menu_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_allowed(update.effective_user.id):
        return ConversationHandler.END

    text = update.message.text

    # ── Список сотрудников ──────────────────────────────────────────────────
    if text == "👥 Список сотрудников":
        users = get_all_users()
        if not users:
            await update.message.reply_text("Пользователей пока нет.", reply_markup=kb_main())
            return MAIN_MENU
        msg = "👥 *Список сотрудников:*\n\n"
        for i, u in enumerate(users, 1):
            msg += f"{i}. {u[1]}\n   📍 {u[2]}  |  📞 {u[3]}\n"
        await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=kb_main())
        return MAIN_MENU

    # ── Отчёт по сотруднику ─────────────────────────────────────────────────
    elif text == "👤 Отчёт по сотруднику":
        kb, users = kb_employees()
        if not users:
            await update.message.reply_text("Сотрудников нет.", reply_markup=kb_main())
            return MAIN_MENU
        context.user_data["users_list"] = users
        await update.message.reply_text("Выберите сотрудника:", reply_markup=kb)
        return CHOOSE_EMPLOYEE

    # ── Сводный отчёт ───────────────────────────────────────────────────────
    elif text == "📊 Сводный отчёт (все)":
        await update.message.reply_text("Выберите период:", reply_markup=kb_period())
        return SUMMARY_PERIOD

    # ── Excel ────────────────────────────────────────────────────────────────
    elif text == "📥 Выгрузка в Excel":
        await update.message.reply_text("Выберите период для выгрузки:", reply_markup=kb_period())
        return EXCEL_PERIOD

    # ── Тур-план ─────────────────────────────────────────────────────────────
    elif text == "📅 Тур-план сотрудника":
        kb, users = kb_employees()
        if not users:
            await update.message.reply_text("Сотрудников нет.", reply_markup=kb_main())
            return MAIN_MENU
        context.user_data["users_list"] = users
        await update.message.reply_text("Выберите сотрудника:", reply_markup=kb)
        return TOURPLAN_EMPLOYEE

    return MAIN_MENU


# ── Отчёт по сотруднику ───────────────────────────────────────────────────────

async def choose_employee(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Главное меню:", reply_markup=kb_main())
        return MAIN_MENU

    users = context.user_data.get("users_list", [])
    chosen = next((u for u in users if f"{u[1]} ({u[2]})" == text), None)
    if not chosen:
        await update.message.reply_text("Сотрудник не найден. Попробуйте снова.")
        return CHOOSE_EMPLOYEE

    context.user_data["chosen_user"] = chosen
    await update.message.reply_text(f"✅ Выбран: {chosen[1]}\nВыберите период:", reply_markup=kb_period())
    return CHOOSE_PERIOD


async def choose_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Главное меню:", reply_markup=kb_main())
        return MAIN_MENU

    if text == "📅 Свободный выбор дат":
        await update.message.reply_text("Введите дату начала (ДД.ММ.ГГГГ):")
        return CHOOSE_START

    start_dt, end_dt, period_name = period_bounds(text)
    await _send_employee_report(update, context, start_dt, end_dt, period_name)
    return MAIN_MENU


async def choose_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["custom_start"] = datetime.strptime(update.message.text, "%d.%m.%Y")
        await update.message.reply_text("Введите дату окончания (ДД.ММ.ГГГГ):")
        return CHOOSE_END
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ. Попробуйте снова:")
        return CHOOSE_START


async def choose_end(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        end_dt   = datetime.strptime(update.message.text, "%d.%m.%Y")
        start_dt = context.user_data["custom_start"]
        if end_dt < start_dt:
            await update.message.reply_text("❌ Дата окончания раньше начала. Попробуйте снова:")
            return CHOOSE_END
        end_dt = end_dt.replace(hour=23, minute=59, second=59)
        await _send_employee_report(update, context, start_dt, end_dt, "период")
        return MAIN_MENU
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ. Попробуйте снова:")
        return CHOOSE_END


async def _send_employee_report(update, context, start_dt, end_dt, period_name):
    user   = context.user_data["chosen_user"]
    visits = get_visits(user[0], start_dt, end_dt)
    await update.message.reply_text(f"⏳ Формирую отчёт по {user[1]}...")

    text     = build_report_text(user, visits, period_name, start_dt, end_dt)
    filename = os.path.join(REPORTS_DIR, f"report_{user[0]}_{datetime.now().strftime('%d%m%Y%H%M%S')}.pdf")
    generate_pdf(text, filename)

    with open(filename, "rb") as f:
        await update.message.reply_document(f, caption=f"📄 Отчёт: {user[1]} за {period_name}")
    await update.message.reply_text("Главное меню:", reply_markup=kb_main())


# ── Сводный отчёт ─────────────────────────────────────────────────────────────

async def summary_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Главное меню:", reply_markup=kb_main())
        return MAIN_MENU
    if text == "📅 Свободный выбор дат":
        context.user_data["summary_mode"] = True
        await update.message.reply_text("Введите дату начала (ДД.ММ.ГГГГ):")
        return SUMMARY_START

    start_dt, end_dt, period_name = period_bounds(text)
    await _send_summary_report(update, context, start_dt, end_dt, period_name)
    return MAIN_MENU


async def summary_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["sum_start"] = datetime.strptime(update.message.text, "%d.%m.%Y")
        await update.message.reply_text("Введите дату окончания (ДД.ММ.ГГГГ):")
        return SUMMARY_END
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ.")
        return SUMMARY_START


async def summary_end(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        end_dt   = datetime.strptime(update.message.text, "%d.%m.%Y").replace(hour=23, minute=59, second=59)
        start_dt = context.user_data["sum_start"]
        if end_dt < start_dt:
            await update.message.reply_text("❌ Дата окончания раньше начала.")
            return SUMMARY_END
        await _send_summary_report(update, context, start_dt, end_dt, "период")
        return MAIN_MENU
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ.")
        return SUMMARY_END


async def _send_summary_report(update, context, start_dt, end_dt, period_name):
    users  = get_all_users()
    visits = get_all_visits(start_dt, end_dt)
    await update.message.reply_text("⏳ Формирую сводный отчёт...")

    # Группируем визиты по сотрудникам
    by_user = {}
    for v in visits:
        by_user.setdefault(v[1], []).append(v)

    full_text = f"📊 СВОДНЫЙ ОТЧЁТ за {period_name}\n"
    full_text += f"Период: {start_dt.strftime('%d.%m.%Y')} — {end_dt.strftime('%d.%m.%Y')}\n"
    full_text += f"Сотрудников: {len(users)}\n"
    full_text += "=" * 40 + "\n"

    user_map = {u[0]: u for u in users}
    for uid, uvists in by_user.items():
        uinfo = user_map.get(uid, (uid, "—", "—", "—"))
        full_text += f"\n👤 {uinfo[1]}  ({uinfo[2]})\n"
        full_text += build_report_text(uinfo, uvists, period_name, start_dt, end_dt)
        full_text += "\n" + "=" * 40 + "\n"

    filename = os.path.join(REPORTS_DIR, f"summary_{datetime.now().strftime('%d%m%Y%H%M%S')}.pdf")
    generate_pdf(full_text, filename)

    with open(filename, "rb") as f:
        await update.message.reply_document(f, caption=f"📊 Сводный отчёт за {period_name}")
    await update.message.reply_text("Главное меню:", reply_markup=kb_main())


# ── Excel выгрузка ────────────────────────────────────────────────────────────

async def excel_period(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Главное меню:", reply_markup=kb_main())
        return MAIN_MENU
    if text == "📅 Свободный выбор дат":
        await update.message.reply_text("Введите дату начала (ДД.ММ.ГГГГ):")
        return EXCEL_START

    start_dt, end_dt, period_name = period_bounds(text)
    await _send_excel(update, context, start_dt, end_dt, period_name)
    return MAIN_MENU


async def excel_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        context.user_data["excel_start"] = datetime.strptime(update.message.text, "%d.%m.%Y")
        await update.message.reply_text("Введите дату окончания (ДД.ММ.ГГГГ):")
        return EXCEL_END
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ.")
        return EXCEL_START


async def excel_end(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        end_dt   = datetime.strptime(update.message.text, "%d.%m.%Y").replace(hour=23, minute=59, second=59)
        start_dt = context.user_data["excel_start"]
        if end_dt < start_dt:
            await update.message.reply_text("❌ Дата окончания раньше начала.")
            return EXCEL_END
        await _send_excel(update, context, start_dt, end_dt, "период")
        return MAIN_MENU
    except ValueError:
        await update.message.reply_text("❌ Формат: ДД.ММ.ГГГГ.")
        return EXCEL_END


async def _send_excel(update, context, start_dt, end_dt, period_name):
    users  = get_all_users()
    visits = get_all_visits(start_dt, end_dt)
    await update.message.reply_text("⏳ Формирую Excel файл...")

    by_user = {}
    for v in visits:
        by_user.setdefault(v[1], []).append(v)

    filename = os.path.join(REPORTS_DIR, f"excel_{datetime.now().strftime('%d%m%Y%H%M%S')}.xlsx")
    generate_excel(users, by_user, period_name, start_dt, end_dt, filename)

    with open(filename, "rb") as f:
        await update.message.reply_document(
            f,
            filename=f"Отчёт_{start_dt.strftime('%d.%m.%Y')}—{end_dt.strftime('%d.%m.%Y')}.xlsx",
            caption=f"📥 Excel-выгрузка за {period_name}"
        )
    await update.message.reply_text("Главное меню:", reply_markup=kb_main())


# ── Тур-план сотрудника ───────────────────────────────────────────────────────

async def tourplan_employee(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    if text == "🔙 Назад":
        await update.message.reply_text("Главное меню:", reply_markup=kb_main())
        return MAIN_MENU

    users   = context.user_data.get("users_list", [])
    chosen  = next((u for u in users if f"{u[1]} ({u[2]})" == text), None)
    if not chosen:
        await update.message.reply_text("Сотрудник не найден. Попробуйте снова.")
        return TOURPLAN_EMPLOYEE

    plan = get_tourplan(chosen[0])
    if not plan:
        await update.message.reply_text(
            f"📅 У {chosen[1]} нет записей в тур-плане.",
            reply_markup=kb_main()
        )
        return MAIN_MENU

    msg = f"📅 *Тур-план: {chosen[1]}*\n\n"
    for row in plan:
        msg += f"📆 {row[1]}  |  📍 {row[2]}\n🏢 {row[3]}\n🎯 {row[4]}\n\n"

    await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=kb_main())
    return MAIN_MENU


# ─── ЗАПУСК ───────────────────────────────────────────────────────────────────

def main():
    app = Application.builder().token(ADMIN_BOT_TOKEN).read_timeout(120).write_timeout(120).build()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            MAIN_MENU:        [MessageHandler(filters.TEXT & ~filters.COMMAND, main_menu_handler)],
            CHOOSE_EMPLOYEE:  [MessageHandler(filters.TEXT & ~filters.COMMAND, choose_employee)],
            CHOOSE_PERIOD:    [MessageHandler(filters.TEXT & ~filters.COMMAND, choose_period)],
            CHOOSE_START:     [MessageHandler(filters.TEXT & ~filters.COMMAND, choose_start)],
            CHOOSE_END:       [MessageHandler(filters.TEXT & ~filters.COMMAND, choose_end)],
            SUMMARY_PERIOD:   [MessageHandler(filters.TEXT & ~filters.COMMAND, summary_period)],
            SUMMARY_START:    [MessageHandler(filters.TEXT & ~filters.COMMAND, summary_start)],
            SUMMARY_END:      [MessageHandler(filters.TEXT & ~filters.COMMAND, summary_end)],
            EXCEL_PERIOD:     [MessageHandler(filters.TEXT & ~filters.COMMAND, excel_period)],
            EXCEL_START:      [MessageHandler(filters.TEXT & ~filters.COMMAND, excel_start)],
            EXCEL_END:        [MessageHandler(filters.TEXT & ~filters.COMMAND, excel_end)],
            TOURPLAN_EMPLOYEE:[MessageHandler(filters.TEXT & ~filters.COMMAND, tourplan_employee)],
        },
        fallbacks=[CommandHandler("start", start)],
    )

    app.add_handler(conv)
    print("✅ Админ-бот запущен")
    app.run_polling()


if __name__ == "__main__":
    main()
